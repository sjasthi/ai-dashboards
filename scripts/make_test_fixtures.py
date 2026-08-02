"""Generate the spreadsheet fixtures that the file-loading tests assert against.

    python scripts/make_test_fixtures.py

Output goes to tests/fixtures/excel/ and is committed, so `pytest` works on a
fresh clone without running this first. Regenerate after editing SPECS, then
re-check tests/fixtures/manifest.json against the new shapes.

Every workbook is written twice - once as .xlsx via openpyxl, once as .xls via
xlwt - because the two formats are read by entirely separate code. .xlsx goes
through workbook_probe._probe_xlsx (openpyxl), .xls through _probe_xls (xlrd),
and they share only _extent, which has to treat an empty cell as None for one
reader and '' for the other. A case that exists only as .xlsx leaves half of
that divergence untested.

xlwt is unmaintained and is not Excel, so it is not trusted for cases that turn
on real Excel write artifacts - notably rows holding formatting but no values,
where the dimension record Excel writes is the whole point. That case uses a
committed, Excel-saved workbook instead.
"""

import csv
import os
import sys
from datetime import datetime

# Fixed so regenerating produces identical bytes and git stays quiet. openpyxl
# otherwise stamps the current time into docProps/core.xml, which would make
# every regeneration look like a content change.
FIXED_TIME = datetime(2026, 1, 1, 0, 0, 0)

# The same instant as a zipfile date_time tuple. Pinning core.xml alone is not
# enough for .xlsx: it is a ZIP, and every member's header carries its own
# mtime, which zipfile takes from the clock. Without normalising those, all six
# .xlsx fixtures show up as modified on every regeneration even though their
# contents are identical. .xls (a flat BIFF stream) and .csv have no such field
# and are already reproducible.
FIXED_ZIP_TIME = (2026, 1, 1, 0, 0, 0)

_FIXTURES_ROOT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "tests", "fixtures",
)
FIXTURE_DIR = os.path.join(_FIXTURES_ROOT, "excel")
CSV_DIR = os.path.join(_FIXTURES_ROOT, "csv")
BROKEN_DIR = os.path.join(_FIXTURES_ROOT, "broken")

# stem -> {sheet name: (data row count, column names)}
#
# Data rows exclude the header. Keep every sheet non-empty unless a case is
# specifically about empty-sheet handling: DataLoader._add_excel drops empty
# sheets, and a dropped sheet also changes whether its siblings get the
# parenthesised "<sheet> (<stem>)<ext>" name, which only applies when more than
# one sheet survives.
SPECS = {
    # --- Required matrix cell 1: one workbook, one sheet -------------------
    "m1_one_book_one_sheet": {
        "Sales": (4, ["id", "product", "units", "revenue"]),
    },

    # --- Required matrix cell 2: two workbooks, one sheet each -------------
    "m2a_one_sheet_orders": {
        "Orders": (3, ["id", "customer", "amount"]),
    },
    "m2b_one_sheet_items": {
        "Items": (5, ["id", "product", "units"]),
    },

    # --- Required matrix cell 3: one workbook, several sheets --------------
    "m3_one_book_many_sheets": {
        "Sales": (4, ["id", "product", "units", "revenue"]),
        "Regions": (3, ["id", "region", "units"]),
        "Staff": (2, ["id", "name", "amount"]),
    },

    # --- Required matrix cell 4: two workbooks, several sheets each --------
    # Sheet names are distinct across the two books here so this stays the
    # clean baseline; the deliberate same-name-in-two-books collision is a
    # separate case.
    "m4a_many_sheets_east": {
        "Orders": (3, ["id", "customer", "amount"]),
        "Returns": (2, ["id", "product", "units"]),
    },
    "m4b_many_sheets_west": {
        "Shipments": (4, ["id", "carrier", "units"]),
        "Inventory": (2, ["id", "product", "units"]),
    },
}


# stem -> (data row count, column names). Plain utf-8, one table per file, for
# the .csv-alone and mixed-batch cases. Encoding variants (BOM, utf-16, cp1252)
# are separate fixtures: they exercise the fallback ladder in
# DataLoader._read_csv_with_encoding rather than table naming.
CSV_SPECS = {
    "csv_orders": (3, ["id", "customer", "amount"]),
    "csv_regions": (5, ["id", "region", "units"]),
}


def sheet_values(n_rows, columns):
    """Header row followed by n_rows of deterministic data.

    Values are derived from the column name so a mis-parsed column is obvious
    when reading a failure message, and from the row index so row counts can be
    eyeballed.
    """
    yield list(columns)
    for i in range(1, n_rows + 1):
        row = []
        for name in columns:
            if name == "id":
                row.append(i)
            elif name in ("units", "qty"):
                row.append(i * 2)
            elif name in ("revenue", "amount"):
                row.append(round(i * 10.5, 2))
            else:
                row.append(f"{name}-{i}")
        yield row


def write_xlsx(path, sheets):
    from openpyxl import Workbook

    wb = Workbook()
    # A new Workbook already has one sheet; drop it so sheet order is exactly
    # the SPECS order rather than "Sheet" first.
    wb.remove(wb.active)
    for name, (n_rows, columns) in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in sheet_values(n_rows, columns):
            ws.append(row)
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    wb.save(path)
    _normalize_xlsx(path)


def _normalize_xlsx(path):
    """Rewrite an .xlsx in place so its bytes depend only on its contents.

    Two independent sources of per-run drift, both of which have to go or the
    fixture churns on every regeneration:

    1. ZIP member mtimes, which zipfile takes from the clock.
    2. <dcterms:modified> in docProps/core.xml. Setting wb.properties.modified
       before save() is not enough -- openpyxl overwrites it with the current
       time while saving, so it has to be patched afterwards.

    Members are copied in their original order with their original compression,
    so nothing but those timestamps changes and the result stays a valid
    workbook -- which the generator verifies by reading every file back.
    """
    import re
    import shutil
    import tempfile
    import zipfile

    stamp = FIXED_TIME.strftime("%Y-%m-%dT%H:%M:%SZ").encode()

    def pin_core_xml(data):
        for field in (b"created", b"modified"):
            data = re.sub(
                rb"(<dcterms:" + field + rb"[^>]*>)[^<]*(</dcterms:" + field + rb">)",
                rb"\g<1>" + stamp + rb"\g<2>",
                data,
            )
        return data

    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(path))
    os.close(fd)
    try:
        with zipfile.ZipFile(path) as src, zipfile.ZipFile(tmp_path, "w") as dst:
            for item in src.infolist():
                data = src.read(item.filename)
                if item.filename == "docProps/core.xml":
                    data = pin_core_xml(data)
                pinned = zipfile.ZipInfo(item.filename, date_time=FIXED_ZIP_TIME)
                pinned.compress_type = item.compress_type
                pinned.external_attr = item.external_attr
                dst.writestr(pinned, data)
        shutil.move(tmp_path, path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def write_xls(path, sheets):
    import xlwt

    wb = xlwt.Workbook()
    for name, (n_rows, columns) in sheets.items():
        ws = wb.add_sheet(name)
        for r, row in enumerate(sheet_values(n_rows, columns)):
            for c, value in enumerate(row):
                ws.write(r, c, value)
    wb.save(path)


def write_csv(path, n_rows, columns, encoding="utf-8"):
    """Write one CSV using the same value generator as the workbooks.

    newline="" plus an explicit lineterminator keeps the bytes identical on
    Windows and POSIX -- csv would otherwise emit \\r\\r\\n here, and a fixture
    whose bytes depend on the machine that generated it is not a fixture.
    """
    with open(path, "w", newline="", encoding=encoding) as fh:
        writer = csv.writer(fh, lineterminator="\n")
        for row in sheet_values(n_rows, columns):
            writer.writerow(row)


# --------------------------------------------------------------------------
# Hazard workbooks. Each is bespoke enough that the SPECS table cannot express
# it, so each gets a builder. Sheet contents matter here in a way row counts
# alone do not: the point is the *names* and the shape of the edges.
#
# Every builder returns {sheet name: [row, ...]} with the header first, so the
# same structure can be written by both openpyxl and xlwt.
# --------------------------------------------------------------------------

# Excel's own hard limit on a sheet name. xlwt enforces it too, so a name any
# longer would make the .xls half of the pair unbuildable rather than testing
# anything about our code.
LONG_SHEET_NAME = "Q1_regional_breakdown_by_store"  # exactly 31 chars


def _rows(n_rows, columns):
    return list(sheet_values(n_rows, columns))


def hazard_sheets():
    """stem -> {sheet name: [rows]}, for the naming and structure cases."""
    return {
        # Parentheses in the sheet name make "<sheet> (<stem>)<ext>" ambiguous:
        # "Items (raw) (hz_parens).xlsx" cannot be parsed back into its parts.
        # DataLoader.origins is the documented mitigation, so these cases exist
        # to assert origins still maps each table to the right workbook.
        "hz_parens": {
            "Items (raw)": _rows(3, ["id", "product", "units"]),
            "Items (clean)": _rows(2, ["id", "product", "units"]),
        },

        # The same sheet names in two different workbooks. Uniqueness has to
        # come from the workbook stem or one table silently overwrites the other
        # in the tables() dict.
        "hz_same_names_a": {
            "Data": _rows(3, ["id", "product", "units"]),
            "Meta": _rows(2, ["id", "name", "amount"]),
        },
        "hz_same_names_b": {
            "Data": _rows(4, ["id", "product", "units"]),
            "Meta": _rows(5, ["id", "name", "amount"]),
        },

        "hz_long_name": {
            LONG_SHEET_NAME: _rows(3, ["id", "region", "units"]),
            "Short": _rows(2, ["id", "name", "amount"]),
        },

        # Non-ASCII in both sheet names and headers. Excel stores these as
        # UTF-16 internally; xlwt writes them as BIFF8 unicode strings.
        "hz_unicode": {
            "Ventas Año": _rows(3, ["id", "región", "unidades"]),
            "Café München": _rows(2, ["id", "produit", "coût"]),
        },

        # Leading/trailing whitespace in headers is stripped
        # (data_loader.py:60), so the loaded columns must come back clean.
        "hz_ws_headers": {
            "Padded": [["  id  ", " product", "units "],
                       [1, "product-1", 2],
                       [2, "product-2", 4]],
            "Second": _rows(2, ["id", "name", "amount"]),
        },

        # Numeric and blank headers must survive untouched: the strip is guarded
        # by isinstance(c, str) precisely so a numeric header does not raise.
        "hz_odd_headers": {
            "Odd": [[2024, None, "note"],
                    [1, "x", "ok"],
                    [2, "y", "ok"]],
            "Second": _rows(2, ["id", "name", "amount"]),
        },

        # A completely blank sheet between populated ones. DataLoader drops it,
        # which also decides whether the survivors get parenthesised names.
        "hz_blank_middle": {
            "First": _rows(3, ["id", "product", "units"]),
            "Blank": [],
            "Last": _rows(2, ["id", "name", "amount"]),
        },

        # Header row but no data. inspect_file reports empty:true and
        # DataLoader skips it -- leaving exactly ONE surviving sheet, so the
        # table name collapses to the bare filename.
        "hz_header_only": {
            "HeaderOnly": [["id", "product", "units"]],
            "Real": _rows(3, ["id", "product", "units"]),
        },

        # Smallest non-empty table: one column, one data row.
        "hz_single_cell": {
            "Tiny": [["id"], [1]],
        },

        # A cell holding only spaces counts as data (workbook_probe.py:38), so
        # the extent must include this row rather than treating it as blank.
        "hz_ws_cell": {
            "Spaces": [["id", "note"], [1, "   "], [2, "real"]],
        },
    }


def write_xlsx_rows(path, sheets):
    """Write explicit row lists (rather than a SPECS entry) to .xlsx."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    wb.save(path)
    _normalize_xlsx(path)


def write_xls_rows(path, sheets):
    """Write explicit row lists to .xls.

    xlwt has no append(); blank cells are simply never written, which is what
    makes the None in hz_odd_headers land as a genuinely empty cell -- the same
    condition workbook_probe._extent has to read as '' from xlrd and None from
    openpyxl.
    """
    import xlwt

    wb = xlwt.Workbook()
    for name, rows in sheets.items():
        ws = wb.add_sheet(name)
        for r, row in enumerate(rows):
            for c, value in enumerate(row):
                if value is not None:
                    ws.write(r, c, value)
        if not rows:
            # xlwt refuses to save a sheet with no cells at all, so a truly
            # blank sheet needs one written-then-blanked cell to exist.
            ws.write(0, 0, "")
    wb.save(path)


def write_formatting_only_trailing(path):
    """A workbook whose used range extends past its data, via formatting alone.

    This is the failure workbook_probe's docstring exists for: rows 5-14 carry a
    fill but no values, so the sheet's dimension reports 14 rows while only 3
    hold data. A naive probe reads 13 data rows and the UI promises rows the
    analysis will never load.

    .xlsx only. In .xls the row count comes from xlrd's own view of the BIFF
    rows, and a fixture built by xlwt would be testing xlwt's idea of a used
    range rather than Excel's.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    for row in _rows(3, ["id", "product", "units"]):
        ws.append(row)
    for r in range(5, 15):
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="FFFF00")
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    wb.save(path)
    _normalize_xlsx(path)


# --------------------------------------------------------------------------
# CSV encodings. The ladder in DataLoader._read_csv_with_encoding is
# ['utf-8', 'utf-16', 'latin-1', 'iso-8859-1', 'cp1252'], tried in order.
#
# enc_cp1252 deliberately uses characters that exist in cp1252 but NOT in
# latin-1 (U+2019 and U+20AC live at 0x92 and 0x80, which are control codes in
# latin-1). Because latin-1 decodes any byte without raising, it always wins
# before cp1252 is reached, and those bytes come back as control characters.
# The manifest records the correct expectation as an expected failure.
# --------------------------------------------------------------------------
ENCODING_SPECS = {
    "enc_utf8": ("utf-8", ["id", "café", "note"]),
    "enc_utf8_bom": ("utf-8-sig", ["id", "café", "note"]),
    "enc_utf16": ("utf-16", ["id", "café", "note"]),
    "enc_cp1252": ("cp1252", ["id", "pr’ce", "cost€"]),
}


def write_broken_fixtures(broken_dir, source_xlsx):
    """Files that must fail, or fall back, without ever returning a 500."""
    zero = os.path.join(broken_dir, "zero_byte.xlsx")
    with open(zero, "wb"):
        pass

    # A real CSV wearing an .xlsx extension. pd.read_excel raises ValueError and
    # DataLoader falls back to a CSV read (data_loader.py:40).
    with open(os.path.join(broken_dir, "really_csv.xlsx"), "w",
              newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh, lineterminator="\n")
        for row in _rows(3, ["id", "product", "units"]):
            writer.writerow(row)

    # A valid workbook cut off mid-archive. Not a zip any more, so the Excel
    # readers reject it -- and then the CSV fallback accepts the binary
    # remains, which is the behaviour the manifest flags as a defect.
    with open(source_xlsx, "rb") as src:
        head = src.read()[:400]
    with open(os.path.join(broken_dir, "truncated.xlsx"), "wb") as fh:
        fh.write(head)

    return ["zero_byte.xlsx", "really_csv.xlsx", "truncated.xlsx"]


def write_scale_fixture(path, n_rows=50_000, n_cols=12):
    """One large workbook, gitignored and built on demand.

    Not committed: it is several MB and nothing about it is interesting except
    its size. BIFF8 caps at 65,536 rows / 256 columns, so .xls cannot carry a
    bigger version of this and only .xlsx is generated.
    """
    from openpyxl import Workbook

    columns = ["id"] + [f"metric_{i}" for i in range(1, n_cols)]
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Big")
    ws.append(columns)
    for i in range(1, n_rows + 1):
        ws.append([i] + [(i * j) % 997 for j in range(1, n_cols)])
    wb.save(path)
    _normalize_xlsx(path)


def verify():
    """Read every generated workbook back through pandas.

    _normalize_xlsx rewrites the archive after openpyxl has finished with it, so
    this is not a formality: a mistake there would produce a file that still
    looks like a workbook but no longer opens, and the failure would otherwise
    surface much later as a confusing test error.
    """
    import pandas as pd

    for stem in SPECS:
        for ext in (".xlsx", ".xls"):
            path = os.path.join(FIXTURE_DIR, f"{stem}{ext}")
            if not os.path.exists(path):
                continue
            sheets = pd.read_excel(path, sheet_name=None)
            expected = {name: n for name, (n, _) in SPECS[stem].items()}
            actual = {name: len(df) for name, df in sheets.items()}
            if actual != expected:
                raise AssertionError(
                    f"{os.path.basename(path)}: read back {actual}, expected {expected}"
                )

    # Hazard workbooks are checked for openability and sheet names rather than
    # row counts: several deliberately contain sheets pandas reads as empty, so
    # a row-count comparison would encode the very thing under test.
    hazards = hazard_sheets()
    for stem, sheets in hazards.items():
        for ext in (".xlsx", ".xls"):
            path = os.path.join(FIXTURE_DIR, f"{stem}{ext}")
            if not os.path.exists(path):
                continue
            got = list(pd.read_excel(path, sheet_name=None))
            want = list(sheets)
            if got != want:
                raise AssertionError(
                    f"{os.path.basename(path)}: sheets {got}, expected {want}"
                )

    print("verified: every workbook opens with the expected sheet names")


def main():
    os.makedirs(FIXTURE_DIR, exist_ok=True)
    os.makedirs(CSV_DIR, exist_ok=True)

    try:
        import xlwt  # noqa: F401
        xls_available = True
    except ImportError:
        xls_available = False
        print("xlwt is not installed - skipping .xls fixtures. "
              "pip install -r requirements-dev.txt", file=sys.stderr)

    written = 0
    for stem, sheets in SPECS.items():
        xlsx_path = os.path.join(FIXTURE_DIR, f"{stem}.xlsx")
        write_xlsx(xlsx_path, sheets)
        print(f"wrote {os.path.basename(xlsx_path)}")
        written += 1

        if xls_available:
            xls_path = os.path.join(FIXTURE_DIR, f"{stem}.xls")
            write_xls(xls_path, sheets)
            print(f"wrote {os.path.basename(xls_path)}")
            written += 1

    for stem, sheets in hazard_sheets().items():
        xlsx_path = os.path.join(FIXTURE_DIR, f"{stem}.xlsx")
        write_xlsx_rows(xlsx_path, sheets)
        print(f"wrote {os.path.basename(xlsx_path)}")
        written += 1

        if xls_available:
            xls_path = os.path.join(FIXTURE_DIR, f"{stem}.xls")
            write_xls_rows(xls_path, sheets)
            print(f"wrote {os.path.basename(xls_path)}")
            written += 1

    fmt_path = os.path.join(FIXTURE_DIR, "hz_fmt_trailing.xlsx")
    write_formatting_only_trailing(fmt_path)
    print(f"wrote {os.path.basename(fmt_path)}")
    written += 1

    for stem, (n_rows, columns) in CSV_SPECS.items():
        csv_path = os.path.join(CSV_DIR, f"{stem}.csv")
        write_csv(csv_path, n_rows, columns)
        print(f"wrote {os.path.basename(csv_path)}")
        written += 1

    for stem, (encoding, columns) in ENCODING_SPECS.items():
        csv_path = os.path.join(CSV_DIR, f"{stem}.csv")
        write_csv(csv_path, 2, columns, encoding=encoding)
        print(f"wrote {os.path.basename(csv_path)} ({encoding})")
        written += 1

    os.makedirs(BROKEN_DIR, exist_ok=True)
    for name in write_broken_fixtures(BROKEN_DIR, fmt_path):
        print(f"wrote broken/{name}")
        written += 1

    verify()

    if "--scale" in sys.argv:
        scale_path = os.path.join(FIXTURE_DIR, "scale_50k.xlsx")
        print("building the 50k-row scale fixture (gitignored, takes a moment)...")
        write_scale_fixture(scale_path)
        print(f"wrote {os.path.basename(scale_path)}")
        written += 1
    else:
        print("\nskipped scale_50k.xlsx - pass --scale to build it")

    print(f"\n{written} file(s) under {_FIXTURES_ROOT}")


if __name__ == "__main__":
    main()
