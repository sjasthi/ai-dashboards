"""Test 1 - extension coverage.

Does every .csv, .xls and .xlsx file in tests/data/extensions/ read and load
correctly, on its own and in mixed batches?

Cases are discovered from the folder tree, so adding a file to
tests/data/extensions/<ext>/ adds a test with no edit here. What each case asserts
lives in tests/loading_checks.py, shared with scripts/run_test_plan.py so the HTML
report and the suite can never disagree.

Parametrisation is per case, so a bad file shows up as one red row naming that
file rather than a single opaque failure.
"""

import glob
import os

import pytest

from tests import loading_checks
from tests.conftest import EXTENSION_CASES, case_ids, require_corpus

SINGLE_FILE_CASES = [c for c in EXTENSION_CASES if c["group"] == "extensions"]
MIXED_CASES = [c for c in EXTENSION_CASES if c["group"] == "extensions-mixed"]


def test_corpus_is_present():
    require_corpus(EXTENSION_CASES, loading_checks.EXTENSIONS_ROOT)
    assert SINGLE_FILE_CASES, "extensions/ has folders but no supported files"


@pytest.mark.parametrize("case", SINGLE_FILE_CASES, ids=case_ids(SINGLE_FILE_CASES))
def test_single_file_loads(case):
    """One file at a time: it reads, it loads, and the probe's row count matches
    what pandas actually produced."""
    result = loading_checks.run_batch(case)
    assert result["passed"], loading_checks.describe_failures(result)


@pytest.mark.parametrize("case", MIXED_CASES, ids=case_ids(MIXED_CASES))
def test_mixed_batch_loads(case):
    """Several extensions uploaded together, the way a user actually would.

    Worth testing separately from the single-file cases because DataLoader
    dispatches per file inside one shared run - a batch is where a naming
    collision or a dropped table would surface.
    """
    result = loading_checks.run_batch(case)
    assert result["passed"], loading_checks.describe_failures(result)


def test_every_extension_is_represented():
    """Guards the corpus itself: a test plan that claims .csv/.xls/.xlsx coverage
    shouldn't quietly pass because one of the three folders is empty."""
    require_corpus(EXTENSION_CASES, loading_checks.EXTENSIONS_ROOT)
    covered = {c["id"].split("/")[1] for c in SINGLE_FILE_CASES}
    assert covered == {"csv", "xls", "xlsx"}, (
        f"extension coverage is incomplete - found {sorted(covered)}, "
        f"missing {sorted({'csv', 'xls', 'xlsx'} - covered)}"
    )


def _true_last_row(rows):
    """Last 1-based row holding a value, computed here rather than imported.

    Deliberately a re-statement of workbook_probe._extent and not a call into it.
    The test below asks whether the corpus can tell the difference between the
    stored dimension and the real extent; borrowing the function under test would
    make both sides of the comparison move together, and a broken _extent would
    look like a corpus with nothing to find.
    """
    last = 0
    for i, row in enumerate(rows, start=1):
        if any(v is not None and v != "" for v in row):
            last = i
    return last


def test_corpus_covers_formatting_only_trailing_rows():
    """The corpus must contain at least one workbook whose stored dimension
    overstates its real extent.

    This is a coverage assertion about the test data, not about the app. Excel
    writes the dimension record over the *used* range, so rows left holding only
    formatting after their data was deleted still count - the exact failure
    workbook_probe scans cells to avoid ([workbook_probe.py:12-21]). Mutation
    testing showed the earlier corpus could not detect it: every file's stored
    dimension happened to equal its true extent, so making _extent trust the
    stored value broke nothing and the suite stayed green.

    extensions/xlsx/format_only_trailing.xlsx closes that - 4 data rows under a
    dimension of A1:C11, the extra six carrying borders and no values. The test
    names no file, so the property can be satisfied by any workbook; it fails only
    if the corpus loses the property entirely.
    """
    import openpyxl

    require_corpus(EXTENSION_CASES, loading_checks.EXTENSIONS_ROOT)
    pattern = os.path.join(loading_checks.DATA_ROOT, "**", "*.xlsx")
    scanned = 0
    for path in sorted(glob.glob(pattern, recursive=True)):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                stored = ws.max_row or 0
                # Reading rows after max_row is what the probe does too, so this
                # costs one streamed pass per sheet and no more.
                if stored > _true_last_row(ws.iter_rows(values_only=True)):
                    return
        finally:
            wb.close()
        scanned += 1

    pytest.fail(
        f"no workbook in the corpus has formatting-only trailing rows ({scanned} "
        ".xlsx files scanned). The probe/loader row-agreement invariant is "
        "therefore untested against the bug it exists to prevent - restore "
        "extensions/xlsx/format_only_trailing.xlsx or an equivalent file."
    )


@pytest.mark.parametrize("case", SINGLE_FILE_CASES, ids=case_ids(SINGLE_FILE_CASES))
def test_matches_recorded_baseline(case, baseline):
    """Once a file's numbers have been reviewed and recorded, they are fixed.

    Files absent from the baseline are invariant-checked only and skip here rather
    than fail - that state is visible as `unrecorded` in the HTML report. A file
    whose bytes have changed reports `stale` and also skips, since the recorded
    numbers describe different content.
    """
    result = loading_checks.run_batch(case)
    if not result["passed"]:
        pytest.skip("invariants failed; see test_single_file_loads for the detail")

    status, drift = loading_checks.baseline_status(result, baseline)
    if status == "unrecorded":
        pytest.skip("not in baseline yet - run scripts/run_test_plan.py --record")
    if status == "stale":
        pytest.skip("file changed since it was recorded - re-run --record")
    assert not drift, (
        f"{case['id']} drifted from its recorded baseline:\n"
        + "\n".join(f"  {d['field']}: recorded {d['expected']!r}, now {d['actual']!r}"
                    for d in drift)
    )
